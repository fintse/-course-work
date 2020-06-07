import numpy as np
import openpyxl


def gauss(A, B, selected):
    for row in range(len(B)):
        print("(", end='')
        for col in range(len(A[row])):
            print("\t{1:10.2f}{0}".format(" " if (selected is None or selected != (row, col)) else "*", A[row][col]),
                  end='')
        print("\t) * (\tX{0}) = (\t{1:10.2f})".format(row + 1, B[row]))


wb = openpyxl.load_workbook('D:\Gauss.xlsx')
sheet = wb.get_sheet_by_name('Лист1')

f = open('D:\gauss.txt', 'w')
row_ex = [1, 6, 14, 20, 28]

for row in row_ex:
    if row == 6 or row == 20 or row == 28:
        myA = np.empty([6, 6], dtype=float)

        for i in range(0, 6):
            for j in range(0, 6):
                myA.itemset((i, j), float(sheet.cell(row=(row + i), column=(j + 1)).value))

        myB = np.empty(6, dtype=float)

        for i in range(0, 6):
            myB.itemset(i, float(sheet.cell(row=(row + i), column=7).value))

    elif row == 14:
        myA = np.empty([4, 4], dtype=float)

        for i in range(0, 4):
            for j in range(0, 4):
                myA.itemset((i, j), float(sheet.cell(row=(row + i), column=(j + 1)).value))

        myB = np.empty(4, dtype=float)

        for i in range(0, 4):
            myB.itemset(i, float(sheet.cell(row=(row + i), column=5).value))
    elif row == 1:
        myA = np.empty([3, 3], dtype=float)

        for i in range(0, 3):
            for j in range(0, 3):
                myA.itemset((i, j), float(sheet.cell(row=(row + i), column=(j + 1)).value))

        myB = np.empty(3, dtype=float)

        for i in range(0, 3):
            myB.itemset(i, float(sheet.cell(row=(row + i), column=4).value))

    print("Исходная система:")
    gauss(myA, myB, None)
    slv = np.linalg.solve(myA, myB)

    print("Решаем:")
    print(slv)
    f.write(str(slv) + '\n')
    print("\n\n")

f.close()
wb.close()